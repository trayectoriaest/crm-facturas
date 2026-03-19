#!/usr/bin/env python3
"""
Trayectoria EST - CRM Facturación Builder
Lee los Excel de facturación desde SharePoint, compila JSX con Node/Babel,
genera index.html con React ya compilado (sin Babel en el browser).

Columnas Excel:
  A: Razón Social
  B: RUT
  C: Fecha Factura
  D: Valor Neto
  E: Valor Neto + IVA
  F: N° Factura
  G: Estado
  H: Plazo de Pago
  I: Fecha de Pago
  J: OC
  K: SII
"""

import os, io, json, subprocess, urllib.request, urllib.parse
from datetime import datetime, date
from openpyxl import load_workbook

DRIVE_ID  = "b!94YSNWupIUmh41_AtdOVSPieLm_WNpBEh9tqCJhq7-HE4RJxxbTATpTpoCXdSMrL"
BASE_PATH = {
    2025: "Administración y Finanzas/2025/Contabilidad EST/Facturación",
    2026: "Administración y Finanzas/2026/Contabilidad/Trayectoria EST/Facturación",
}
FILES     = {
    2025: "Facturación EST 2025.xlsx",
    2026: "Facturación EST 2026.xlsx",
}

COL_RAZON=0; COL_RUT=1; COL_FECHA=2; COL_NETO=3; COL_TOTAL=4
COL_NFACT=5; COL_ESTADO=6; COL_PLAZO=7; COL_FPAGO=8; COL_OC=9; COL_SII=10

def get_token():
    url  = f"https://login.microsoftonline.com/{os.environ['TENANT_ID']}/oauth2/v2.0/token"
    data = urllib.parse.urlencode({"grant_type":"client_credentials","client_id":os.environ["CLIENT_ID"],"client_secret":os.environ["CLIENT_SECRET"],"scope":"https://graph.microsoft.com/.default"}).encode()
    with urllib.request.urlopen(urllib.request.Request(url, data=data)) as r:
        return json.loads(r.read())["access_token"]

def get_file_bytes(token, year, filename):
    path = f"{BASE_PATH[year]}/{filename}"
    url  = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/root:/{urllib.parse.quote(path)}:/content"
    try:
        with urllib.request.urlopen(urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})) as r:
            return r.read()
    except Exception as e:
        print(f"⚠️  No se pudo leer {filename}: {e}")
        return None

def parse_excel(file_bytes, year):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row): continue
        def v(idx): return row[idx] if idx < len(row) else None
        def fd(val):
            if val is None: return ""
            if isinstance(val, (datetime, date)): return val.strftime("%Y-%m-%d")
            return str(val)
        def fn(val):
            try: return float(val) if val not in (None,"") else 0
            except: return 0
        rows.append({"row":i,"year":year,"razon":str(v(COL_RAZON) or "").strip(),"rut":str(v(COL_RUT) or "").strip(),"fecha":fd(v(COL_FECHA)),"neto":fn(v(COL_NETO)),"total":fn(v(COL_TOTAL)),"nFactura":str(v(COL_NFACT) or "").strip(),"estado":str(v(COL_ESTADO) or "").strip(),"plazo":str(v(COL_PLAZO) or "").strip(),"fechaPago":fd(v(COL_FPAGO)),"oc":str(v(COL_OC) or "").strip(),"sii":str(v(COL_SII) or "").strip()})
    return rows

def build_html(facturas):
    today_str = date.today().strftime("%d/%m/%Y")
    fj = json.dumps(facturas, ensure_ascii=False, separators=(',',':'))

    # Read JSX template
    jsx_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'crm_template.jsx')
    with open(jsx_path, 'r', encoding='utf-8') as f:
        jsx = f.read()

    # Prepare JSX for compilation (placeholder data)
    jsx_compile = jsx.replace('__FJ__', '[]').replace('__TODAY__', '"TODAYPH"')
    tmp_jsx = '/tmp/_crm_build.jsx'
    tmp_js  = '/tmp/_babel_run.js'
    with open(tmp_jsx, 'w', encoding='utf-8') as f:
        f.write(jsx_compile)

    babel_script = """
const babel = require('/tmp/babel/node_modules/@babel/core');
const fs = require('fs');
const code = fs.readFileSync('/tmp/_crm_build.jsx', 'utf8');
const result = babel.transformSync(code, {
  presets: ['/tmp/babel/node_modules/@babel/preset-react'],
  filename: 'crm.jsx'
});
process.stdout.write(result.code);
"""
    with open(tmp_js, 'w') as f:
        f.write(babel_script)

    result = subprocess.run(['node', tmp_js], capture_output=True, text=True)
    if result.returncode != 0:
        raise RuntimeError(f"Babel compilation failed: {result.stderr[:500]}")

    compiled_js = result.stdout
    compiled_js = compiled_js.replace('const FACTURAS_RAW = [];', f'const FACTURAS_RAW = {fj};', 1)
    compiled_js = compiled_js.replace('"TODAYPH"', f'"{today_str}"')

    css = """*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
    :root{--bg:#f0f4fa;--card:#fff;--border:#dde4f0;--border-lt:#eef2f9;--navy:#1e2d6e;--blue:#3a6fd8;--blue-lt:#5fb3f0;--accent:#3a6fd8;--accent-hv:#2d5bbf;--accent-bg:#eef4fd;--danger:#d93b3b;--danger-bg:#fff0f0;--success:#1a9e6a;--success-bg:#edfaf4;--warn:#d97b10;--warn-bg:#fff7ed;--text:#1a2340;--text-2:#3d4f72;--muted:#7d90b5;--radius:10px;--shadow:0 1px 4px rgba(30,45,110,.08);--shadow-md:0 4px 16px rgba(30,45,110,.10)}
    body{background:var(--bg);color:var(--text);font-family:'Inter',sans-serif;font-size:14px}
    .crm-root{display:grid;grid-template-columns:240px 1fr;min-height:100vh}
    .sidebar{background:var(--navy);display:flex;flex-direction:column;position:sticky;top:0;height:100vh}
    .sidebar-logo{padding:24px 20px 20px;border-bottom:1px solid rgba(255,255,255,.1);margin-bottom:8px}
    .logo-mark{display:flex;align-items:center;gap:10px}
    .logo-icon{width:34px;height:34px;flex-shrink:0}.logo-icon svg{width:100%;height:100%}
    .logo-text{color:#fff;font-size:1.05rem;font-weight:700;line-height:1.2}
    .logo-text span{display:block;font-size:.7rem;font-weight:400;color:rgba(255,255,255,.5);margin-top:1px}
    .nav-section{padding:8px 12px;font-size:.65rem;color:rgba(255,255,255,.35);text-transform:uppercase;letter-spacing:1.2px;margin-top:4px}
    .nav-btn{display:flex;align-items:center;gap:10px;padding:10px 14px;margin:2px 8px;border-radius:8px;border:none;background:transparent;color:rgba(255,255,255,.65);font-family:'Inter',sans-serif;font-size:.84rem;font-weight:500;cursor:pointer;transition:all .15s;text-align:left;width:calc(100% - 16px)}
    .nav-btn:hover{background:rgba(255,255,255,.08);color:#fff}
    .nav-btn.active{background:var(--blue);color:#fff;box-shadow:0 2px 8px rgba(58,111,216,.4)}
    .nav-icon{font-size:.95rem;width:20px;text-align:center;flex-shrink:0}
    .nav-badge{margin-left:auto;background:var(--danger);color:#fff;font-size:.65rem;font-weight:700;padding:2px 7px;border-radius:20px}
    .sync-info{margin-top:auto;padding:16px;border-top:1px solid rgba(255,255,255,.1);font-size:.7rem;color:rgba(255,255,255,.35);text-align:center}
    .main{padding:28px 32px;overflow-y:auto;background:var(--bg)}
    .page-header{margin-bottom:24px}
    .page-title{font-size:1.5rem;font-weight:700;color:var(--navy);margin-bottom:4px}
    .page-sub{color:var(--muted);font-size:.82rem}
    .stats-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:24px}
    .stat-card{background:var(--card);border:1px solid var(--border);border-radius:var(--radius);padding:18px 20px;box-shadow:var(--shadow)}
    .stat-card-top{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:10px}
    .stat-label{font-size:.72rem;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.7px}
    .stat-icon{width:34px;height:34px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:1rem}
    .stat-val{font-size:1.45rem;font-weight:700;color:var(--navy)}
    .stat-sub{font-size:.72rem;color:var(--muted);margin-top:3px}
    .table-wrap{background:var(--card);border:1px solid var(--border);border-radius:var(--radius);overflow:hidden;margin-bottom:20px;box-shadow:var(--shadow)}
    .table-header{display:flex;align-items:center;justify-content:space-between;padding:14px 18px;border-bottom:1px solid var(--border-lt);flex-wrap:wrap;gap:10px}
    .table-title{font-weight:600;font-size:.88rem;color:var(--navy)}
    table{width:100%;border-collapse:collapse}
    thead tr{background:#f7f9fd}
    th{padding:10px 14px;text-align:left;font-size:.7rem;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.8px;white-space:nowrap;border-bottom:1px solid var(--border)}
    td{padding:11px 14px;font-size:.83rem;border-top:1px solid var(--border-lt);color:var(--text-2)}
    tr:hover td{background:#f8faff}
    .badge{display:inline-flex;align-items:center;padding:3px 10px;border-radius:20px;font-size:.7rem;font-weight:600;white-space:nowrap}
    .badge-danger{background:var(--danger-bg);color:var(--danger)}.badge-warn{background:var(--warn-bg);color:var(--warn)}.badge-success{background:var(--success-bg);color:var(--success)}.badge-muted{background:#f0f2f7;color:var(--muted)}
    .btn{padding:8px 16px;border-radius:8px;border:none;font-family:'Inter',sans-serif;font-size:.82rem;font-weight:600;cursor:pointer;transition:all .15s;display:inline-flex;align-items:center;gap:6px}
    .btn-primary{background:var(--accent);color:#fff}.btn-primary:hover{background:var(--accent-hv)}
    .btn-ghost{background:transparent;color:var(--text-2);border:1px solid var(--border)}.btn-ghost:hover{border-color:var(--blue);color:var(--blue)}
    .btn-success{background:var(--success-bg);color:var(--success);border:1px solid rgba(26,158,106,.2)}.btn-success:hover{background:#d5f5ea}
    .btn-sm{padding:5px 11px;font-size:.76rem}.btn:disabled{opacity:.45;cursor:not-allowed}
    .search-bar{background:#f5f7fd;border:1px solid var(--border);border-radius:8px;padding:8px 14px;color:var(--text);font-family:'Inter',sans-serif;font-size:.83rem;outline:none;width:220px}
    .search-bar:focus{border-color:var(--blue);background:#fff}
    .select-bar{background:#f5f7fd;border:1px solid var(--border);border-radius:8px;padding:8px 14px;color:var(--text);font-family:'Inter',sans-serif;font-size:.83rem;outline:none}
    .filter-bar{display:flex;gap:8px;margin-bottom:16px;flex-wrap:wrap;align-items:center}
    .empty-state{text-align:center;padding:44px;color:var(--muted)}
    .progress-bar{height:6px;background:var(--border);border-radius:4px;overflow:hidden}.progress-fill{height:100%;border-radius:4px;background:linear-gradient(90deg,var(--blue-lt),var(--blue))}
    .two-col{display:grid;grid-template-columns:1fr 1fr;gap:20px}
    .overlay{position:fixed;inset:0;background:rgba(20,30,70,.45);display:flex;align-items:center;justify-content:center;z-index:200;backdrop-filter:blur(2px)}
    .modal{background:#fff;border:1px solid var(--border);border-radius:14px;padding:28px;width:520px;max-height:92vh;overflow-y:auto;box-shadow:var(--shadow-md)}
    .modal-title{font-size:1.1rem;font-weight:700;color:var(--navy);margin-bottom:16px}
    .modal-actions{display:flex;gap:10px;justify-content:flex-end;margin-top:20px;padding-top:14px;border-top:1px solid var(--border-lt)}
    .kv-grid{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:16px}
    .kv{display:flex;flex-direction:column;gap:3px;background:#f7f9fd;border-radius:8px;padding:10px 12px}
    .kv-key{font-size:.69rem;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.5px}
    .kv-val{font-size:.88rem;font-weight:600;color:var(--navy)}
    .toast{position:fixed;bottom:24px;right:24px;background:var(--navy);color:#fff;padding:12px 20px;border-radius:10px;font-size:.84rem;font-weight:600;z-index:999;display:flex;align-items:center;gap:8px}
    .atraso-chip{display:inline-block;padding:2px 9px;border-radius:6px;font-size:.7rem;font-weight:600}"""

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <title>CRM Facturación — Trayectoria EST</title>
  <script src="https://cdnjs.cloudflare.com/ajax/libs/react/18.2.0/umd/react.production.min.js"></script>
  <script src="https://cdnjs.cloudflare.com/ajax/libs/react-dom/18.2.0/umd/react-dom.production.min.js"></script>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet"/>
  <style>{css}</style>
</head>
<body>
<div id="root"></div>
<script>
{compiled_js}
</script>
</body>
</html>"""

def main():
    print("🔐 Obteniendo token Microsoft Graph...")
    token = get_token()
    all_facturas = []
    for year, filename in FILES.items():
        print(f"📥 Leyendo {filename}...")
        data = get_file_bytes(token, year, filename)
        if data:
            rows = parse_excel(data, year)
            all_facturas.extend(rows)
            print(f"   ✅ {len(rows)} filas leídas")
        else:
            print(f"   ⚠️  Skipping {filename}")
    print(f"\n📊 Total: {len(all_facturas)} facturas")
    print("🏗️  Compilando JSX con Babel...")
    html = build_html(all_facturas)
    with open("index.html", "w", encoding="utf-8") as f:
        f.write(html)
    print("✅ index.html generado correctamente")

if __name__ == "__main__":
    main()
